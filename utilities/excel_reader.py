from openpyxl import load_workbook


class ExcelReader:

    @staticmethod
    def get_data(file_path, sheet_name):

        workbook = load_workbook(file_path)
        print(file_path)
        print(workbook.sheetnames)

        sheet = workbook[sheet_name]

        data = []

        rows = sheet.max_row
        cols = sheet.max_column

        for r in range(2, rows + 1):

            row_data = []

            for c in range(1, cols + 1):

                row_data.append(sheet.cell(r, c).value)

            data.append(row_data)

        workbook.close()

        return data